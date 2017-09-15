from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from re import split

import config.user_info as usr
from config.gsheets import client

class Form:
    """
    Represents an Excel form template, that can additionally
    be filled with data.
    """

    def __init__(self, filepath=None):
        self.filepath = filepath

    def _check_filepath(self):
        if self.filepath == None:
            raise ValueError("filepath must be specified")

    def _save_file(self, wb, filepath=None):
        """
        Saves file if filepath is specified.
        Defaults to using self.filepath if no other filepath is specified.
        Doesn't save if filepath isn't specified.
        """
        if filepath == None:
            filepath = self.filepath
        if filepath:
            wb.save(filepath)

    def _calculate_fund_total(self, ws):
        """
        Given a workbook form with all data already filled in, will
        additionally calculate fund totals.
        """
        amt_cell = usr.tmp_vars['evt_amt']
        cat_cell = usr.tmp_vars['evt_fund']
        for fund in usr.tmp_funds:
            ws[usr.tmp_funds[fund]] = '=sumif({cat},"{fund}",{amt})'.format(
                                       amt=amt_cell,
                                       fund=fund,
                                       cat=cat_cell)

        ws[usr.tmp_funds['TOTAL']] = '=sum({start}:{end})'.format(
                                      start=usr.tmp_funds['SOFC'],
                                      end=usr.tmp_funds['CLCE'])

    def _add_seal(self, ws):
        """
        Given a worksheet template, adds seal.
        """
        # seal image
        seal = Image("temp/wellesley_seal.png", size=(75, 91))
        ws.add_image(seal, usr.tmp_vars['seal'])

    def get_new(self):
        """
        Returns new template with org_name, bookkeeper,
        treasurer and address stub variables filled.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filename=usr.TEMPPATH)
        ws = wb.active

        ws[usr.tmp_vars['org_name']] = usr.ORG_NAME
        ws[usr.tmp_vars['bookkeeper']] = usr.BOOKKEEPER
        ws[usr.tmp_vars['treasurer']] = usr.TREASURER
        ws[usr.tmp_vars['address']] = usr.ADDRESS
        ws[usr.tmp_vars['sofc_fund_num']] = usr.SOFC_FUND_NUM

        self._add_seal(ws)
        self._save_file(wb)
        return wb

    def get_completed(self, form_data):
        """
        Return completed form given a dictionary of values
        to plug into the template.
        Optionally saves as an Excel workbook if filepath is
        specified.

        form_data: Dictionary of user specific data for each form, where
        keys are strings corresponding to usr.tmp_vars key names. 
        """
        wb = self.get_new()
        ws = wb.active
        for varname in form_data:
            # only available data will be filled in
            ws[usr.tmp_vars[varname]] = form_data[varname]
        self._calculate_fund_total(ws)
        self._save_file(wb)
        return wb

class StudentsFile:
    """
    Represents an additional text file that lists the names of students
    who were associated with this reimbursement.
    """

    def __init__(self, filepath=None):
        self.filepath = filepath

    def _split_student_usernames(self, usernames):
        """
        usernames (string): a string list of usernames.
        converts this into a list of string usernames for further use.
        """
        return split('\s*,\s*', usernames)

    def _get_students_data(self, usernames, user_data):
        """
        usernames (string): string list of usernames.
        user_data (dictionary): dictionary of dictionaries mapping a students
        username to their contact data.
        returns a list of dictionaries of students and their contact info.
        """
        usernames = self._split_student_usernames(usernames)
        try:
            student_data = [user_data[user] for user in usernames]
        except KeyError:
            error_msg = """A user from below has not submitted contact info.
            List of student usernames: {}""".format(", ".join(usernames))
            raise ValueError(error_msg)
        return student_data

    def get_new_file(self, usernames, user_data):
        """
        usernames (string): string list of usernames.
        user_data (dictionary): dictionary of dictionaries mapping a students
        username to their contact data.
        """
        print "Generating file {}".format(self.filepath)
        file_body = "This reimbursement paid for the following students " + \
                    "(Banner ID included in brackets for each):"
        student_data = self._get_students_data(usernames, user_data)
        
        with open(self.filepath, "w") as students_file:
            students_file.write(file_body)
            for student in student_data:
                line = "\n- {name} ({id})".format(
                    name=student['stu_name'],
                    id=student['stu_id'])
                students_file.write(line)



