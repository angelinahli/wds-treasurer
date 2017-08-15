import gspread
from openpyxl import load_workbook

import config.user_info as usr
from config.gsheets import client

# (1) Use openpyxl to create templates that can be filled with data

class Template:
    """
    Represents an Excel form template, that can additionally
    be filled with data.
    """

    def _save_file(self, wb, filepath):
        """
        Saves file if filepath is specified.
        """
        if filepath:
            wb.save(filepath)

    def get_empty(self, filepath=None):
        """
        Returns empty template.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filepath='template.xlsx')
        self._save_file(wb, filepath)
        return wb

    def get_new(self, filepath=None):
        """
        Returns new template with org_name, bookkeeper,
        treasurer and address stub variables filled.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filepath = 'template.xlsx')
        ws = wb.active

        ws[usr.tmp_vars['org_name']] = usr.org_name
        ws[usr.tmp_vars['bookkeeper']] = usr.bookkeeper
        ws[usr.tmp_vars['treasurer']] = usr.treasurer
        ws[usr.tmp_vars['address']] = usr.address

        self._save_file(wb, filepath)
        return wb

    def get_completed(self, data_dict, filepath=None):
        """
        Return completed form given a dictionary of values
        to plug into the template.
        Optionally saves as an Excel workbook if filepath is
        specified.

        data_dict: Dictionary of user specific data for each form, where
        keys are strings corresponding to usr.tmp_vars key names. 
        """
        wb = self.get_new()
        ws = wb.active
        for varname in data_dict:
            # only available data will be filled in
            ws[usr.tmp_vars[varname]] = data_dict[varname]
        self._save_file(wb, filepath)
        return wb

# (2) get data for each user

def get_user_dict():
    """
    returns dictionary of dictionaries for each user (student member),
    with username as key, and user's dictionary data as values.
    """
    cols = usr.contacts_cols
    contacts = client.open_by_url(usr.contacts_url).sheet1
    user_dict = {}
    for row in range(2, contacts.row_count):
        data = contacts.row_values(row)
        if not any(data):
            break
        user_dict[ data[cols['username'] - 1] ] = {
            'stu_name': data[cols['stu_name'] - 1],
            'stu_id': data[cols['stu_id'] - 1],
            'stu_unit_box': data[cols['stu_unit_box'] - 1]
        }

    return user_dict

def get_new_row_range(ws):
    """
    given a gspread worksheet representing the form accepting new
    reimbursements, returns a list of the row range that have not
    yet been processed.
    """
    row_range = []
    for row in range(2, ws.row_count):
        # if timestamp is missing data assumed to be missing
        if not ws.cell(row, usr.rmbs_cols['date']):
            break
        # has form is None or false if no form has been created yet
        if not ws.cell(row, usr.rmbs_cols['has_form']):
            row_range.append(row)
    return row_range

def update_rows(ws, row_range):
    """
    given a worksheet and a list of row range, updates rows to indicate a 
    form has been created for each of those rows.
    """
    for row in row_range:
        ws.update_cell(row, usr.rmbs_cols['has_form'], True)
    print "rows {} have been processed!".format(row_range)

def get_rmbs_dict(row_data):
    """
    converts a list of reimbursements data for a given row to a dictionary
    of that data.
    """
    new_data = {}
    for var in usr.rmbs_cols:
        new_data[var] = row_data[rmbs_cols[var] - 1]
    return new_data

def get_data_dict(ws, row_number, user_data):
    """
    given a gspread worksheet of the form accepting new reimbursements,
    the row number to access and a dictionary of data for each user,
    returns a dictionary of all data associated with that row of reimbursements.
    """
    rmbs_data = get_rmbs_dict(ws.row_values(row))
    username = rmbs_data[usr.rmbs_cols['username']]
    try:
        user_values = dict(user_data[username], **rmbs_data)
    except KeyError:
        raise ValueError("User {} has not submitted their info to the contacts form".format(username))
    return {var: user_values.get(var, None) for var in usr.tmp_vars}

def get_worksheet_data(ws, row_range):
    """
    given a gspread worksheet representing the form accepting new 
    reimbursements and list of row range, returns a list of dictionaries 
    that can be used to create spreadsheets.
    """
    user_data = get_user_dict()
    return [get_data_dict(ws, row, user_data) for row in row_range]

# (3) make spreadsheets for users and update converted rows

def make_spreadsheets():
    """
    """
    rmbs = client.open_by_url(usr.rmbs_url).sheet1
