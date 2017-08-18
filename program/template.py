from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

import config.user_info as usr
from config.gsheets import client

class FormTemplate:
    """
    Represents an Excel form template, that can additionally
    be filled with data.
    """

    def __init__(self, filepath=None):
        self.filepath = filepath

    def _check_filepath(self):
        if self.filepath == None:
            raise ValueError("filepath must be specified")

    def _save_file(self, wb, filepath=None):
        """
        Saves file if filepath is specified.
        Defaults to using self.filepath if no other filepath is specified.
        Doesn't save if filepath isn't specified.
        """
        if filepath == None:
            filepath = self.filepath
        if filepath:
            wb.save(filepath)

    def _calculate_fund_total(self, ws):
        """
        Given a workbook form with all data already filled in, will
        additionally calculate fund totals.
        """
        amt_cell = usr.tmp_vars['evt_amt']
        cat_cell = usr.tmp_vars['evt_fund']
        for fund in usr.tmp_funds:
            ws[usr.tmp_funds[fund]] = '=sumif({cat},"{fund}",{amt})'.format(
                                       amt=amt_cell,
                                       fund=fund,
                                       cat=cat_cell)

        ws[usr.tmp_funds['TOTAL']] = '=sum({start}:{end})'.format(
                                      start=usr.tmp_funds['SOFC'],
                                      end=usr.tmp_funds['CLCE'])

    def _add_seal(self, ws):
        """
        Given a worksheet template, adds seal.
        """
        # seal image
        seal = Image("temp/wellesley_seal.png", size=(75, 91))
        ws.add_image(seal, usr.tmp_vars['seal'])

    def get_empty(self):
        """
        Returns empty template.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filename=usr.TEMPPATH)
        self._add_seal(ws)
        self._save_file(wb)
        return wb

    def get_new(self):
        """
        Returns new template with org_name, bookkeeper,
        treasurer and address stub variables filled.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filename=usr.TEMPPATH)
        ws = wb.active

        ws[usr.tmp_vars['org_name']] = usr.ORG_NAME
        ws[usr.tmp_vars['bookkeeper']] = usr.BOOKKEEPER
        ws[usr.tmp_vars['treasurer']] = usr.TREASURER
        ws[usr.tmp_vars['address']] = usr.ADDRESS

        self._add_seal(ws)
        self._save_file(wb)
        return wb

    def get_completed(self, data_dict):
        """
        Return completed form given a dictionary of values
        to plug into the template.
        Optionally saves as an Excel workbook if filepath is
        specified.

        data_dict: Dictionary of user specific data for each form, where
        keys are strings corresponding to usr.tmp_vars key names. 
        """
        wb = self.get_new()
        ws = wb.active
        for varname in data_dict:
            # only available data will be filled in
            ws[usr.tmp_vars[varname]] = data_dict[varname]
        self._calculate_fund_total(ws)
        self._save_file(wb)
        return wb

    def get_existing(self):
        """
        retrieves and returns an existing template.
        """
        self._check_filepath()
        wb = load_workbook(filename=self.filepath)
        ws = wb.active
        self._add_seal(ws)
        return wb

    def format_existing(self):
        """
        formats existing template and then returns new file and saves
        with new filename.
        """
        self._check_filepath()
        wb = self.get_existing()
        ws = wb.active
        new_filepath = '{old_filename}_new.xlsx'.format(
                         old_filename=self.filepath.rsplit('.', 1)[0])
        self._add_seal(ws)
        self._calculate_fund_total(ws)
        self._save_file(wb, new_filepath)