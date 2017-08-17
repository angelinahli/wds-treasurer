from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

import config.user_info as usr
from config.gsheets import client

class FormTemplate:
    """
    Represents an Excel form template, that can additionally
    be filled with data.
    """

    def _save_file(self, wb, filepath):
        """
        Saves file if filepath is specified.
        """
        if filepath:
            wb.save(filepath)

    def _calculate_fund_total(self, ws):
        """
        Given a workbook form with all data already filled in, will
        additionally calculate fund totals.
        """
        amt_cell = usr.tmp_vars['evt_amt']
        cat_cell = usr.tmp_vars['evt_fund']
        for fund in ['SOFC', 'PROFITS', 'CLCE']:
            ws[usr.tmp_funds[fund]] = '=sumif({cat},"{fund}",{amt})'.format(
                                       amt=amt_cell,
                                       fund=fund,
                                       cat=cat_cell)

        ws[usr.tmp_funds['TOTAL']] = '=sum({start}:{end})'.format(
                                      start=usr.tmp_funds['SOFC'],
                                      end=usr.tmp_funds['CLCE'])

    def _format_sheet(self, ws):
        """
        Given a worksheet template, adds seal.
        """
        # seal image
        seal = Image("temp_files/wellesley_seal.png", size=(75, 91))
        ws.add_image(seal, usr.tmp_vars['seal'])

    def get_empty(self, filepath=None):
        """
        Returns empty template.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filename='temp_files/template.xlsx')
        self._format_sheet(ws)
        self._save_file(wb, filepath)
        return wb

    def get_new(self, filepath=None):
        """
        Returns new template with org_name, bookkeeper,
        treasurer and address stub variables filled.
        Optionally saves as an Excel workbook if filepath is
        specified.
        """
        wb = load_workbook(filename='temp_files/template.xlsx')
        ws = wb.active

        ws[usr.tmp_vars['org_name']] = usr.ORG_NAME
        ws[usr.tmp_vars['bookkeeper']] = usr.BOOKKEEPER
        ws[usr.tmp_vars['treasurer']] = usr.TREASURER
        ws[usr.tmp_vars['address']] = usr.ADDRESS

        self._format_sheet(ws)
        self._save_file(wb, filepath)
        return wb

    def get_completed(self, data_dict, filepath=None):
        """
        Return completed form given a dictionary of values
        to plug into the template.
        Optionally saves as an Excel workbook if filepath is
        specified.

        data_dict: Dictionary of user specific data for each form, where
        keys are strings corresponding to usr.tmp_vars key names. 
        """
        wb = self.get_new()
        ws = wb.active
        for varname in data_dict:
            # only available data will be filled in
            ws[usr.tmp_vars[varname]] = data_dict[varname]
        self._calculate_fund_total(ws)
        self._save_file(wb, filepath)
        return wb

    def get_existing(self, filepath):
        """
        retrieves and returns an existing template.
        """
        return load_workbook(filename=filepath)

    def format_existing(self, filepath):
        """
        formats existing template and then returns new file and saves
        with new filename.
        """
        wb = self.get_existing(filepath)
        ws = wb.active
        new_filepath = '{old_filename}_new.xlsx'.format(
                         old_filename=filepath.split('.')[0])

        self._format_sheet(ws)
        self._calculate_fund_total(ws)
        self._save_file(wb, new_filepath)
